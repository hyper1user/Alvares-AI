"""
Модуль для заповнення багатомісячного табелю з даних місячних файлів
"""
import os
import openpyxl
from datetime import datetime, timedelta
from typing import List, Dict, Tuple, Optional
from collections import defaultdict
import re
from br_calculator import parse_date_from_excel_cell
from month_utils import get_available_months, parse_month_sheet_name, get_source_filename, MONTH_NAMES_UK_LOWER


class SoldierPeriod:
    """Клас для зберігання інформації про період участі військовослужбовця"""
    
    def __init__(self, pib: str, rank: str = "", position: str = ""):
        self.pib = pib
        self.rank = rank
        self.position = position
        self.periods_100: List[Tuple[datetime, datetime]] = []  # (початок, кінець)
        self.periods_30: List[Tuple[datetime, datetime]] = []
        self.periods_0: List[Tuple[datetime, datetime]] = []


class SourceFileReader:
    """Клас для читання даних з місячних .xlsx файлів"""
    
    def __init__(self, source_file: str):
        self.source_file = source_file
        self.wb = None
        
    def load_workbook(self):
        """Завантажує Excel файл"""
        self.wb = openpyxl.load_workbook(self.source_file, data_only=True)
        print(f"Завантажено файл: {self.source_file}")
        print(f"Листи: {self.wb.sheetnames}")
    
    def parse_soldier_info(self, cell_b_value: str, cell_f_value: str = None) -> Tuple[str, str, str]:
        """
        Розбирає дані з стовпця B на звання, ПІБ, посаду
        Якщо не вдається - використовує стовпець F як fallback
        
        Args:
            cell_b_value: Значення зі стовпця B
            cell_f_value: Значення зі стовпця F (ПІБ окремо)
            
        Returns:
            Tuple[rank, pib, position]
        """
        if not cell_b_value:
            # Якщо є F - використаємо його
            if cell_f_value:
                return ("", cell_f_value.strip(), "")
            return ("", "", "")
        
        cell_b_str = str(cell_b_value).strip()
        
        # Спробуємо розпарсити формат: "звання ПІБ, посада"
        # Шукаємо кому як роздільник
        if "," in cell_b_str:
            parts = cell_b_str.split(",", 1)
            rank_and_pib = parts[0].strip()
            position = parts[1].strip() if len(parts) > 1 else ""
            
            # Спробуємо розділити звання та ПІБ
            # ПІБ зазвичай складається з великих літер, може містити малі (Євген, Іван тощо)
            # Шукаємо послідовність слів, що починаються з великої літери (2-3 слова, виглядають як ПІБ)
            # ПІБ зазвичай розташований в кінці першої частини (перед комою)
            words = rank_and_pib.split()
            # Шукаємо ПІБ - послідовність слів, що починаються з великої кириличної літери
            pib_start_idx = None
            for i, word in enumerate(words):
                if word and word[0] in 'АБВГҐДЕЄЖЗИІЇЙКЛМНОПРСТУФХЦЧШЩЮЯ':
                    # Перевіряємо, чи це виглядає як початок ПІБ (великими літерами або з великої)
                    if word.isupper() or (len(word) > 1 and word[0].isupper()):
                        pib_start_idx = i
                        break
            
            if pib_start_idx is not None:
                # Розділяємо на звання (до ПІБ) та ПІБ (від pib_start_idx)
                rank = " ".join(words[:pib_start_idx]).strip()
                pib = " ".join(words[pib_start_idx:]).strip()
                return (rank, pib, position)
        
        # Якщо не вдалося розпарсити, але є F - використовуємо його
        if cell_f_value:
            return ("", cell_f_value.strip(), "")
        
        # Якщо F немає, спробуємо витягти хоча б ПІБ (великі літери)
        pib_match = re.search(r'([А-ЯЄІЇҐ][А-ЯЄІЇҐ\s]+)', cell_b_str)
        if pib_match:
            pib = pib_match.group(1).strip()
            rank = cell_b_str.replace(pib, "").strip()
            return (rank, pib, "")
        
        # Останній варіант - повертаємо все як ПІБ
        return ("", cell_b_str, "")
    
    def parse_date(self, cell_value, take_first: bool = False) -> Optional[datetime]:
        """
        Парсить дату з комірки Excel
        
        Args:
            cell_value: Значення комірки
            take_first: Якщо True, бере першу дату з рядка (для початку періоду)
                        Якщо False, бере останню дату (для кінця періоду)
        """
        if not cell_value:
            return None
        
        try:
            # Якщо це вже datetime
            if isinstance(cell_value, datetime):
                return cell_value
            
            # Якщо це рядок
            if isinstance(cell_value, str):
                cell_str = cell_value.strip()
                
                # Якщо в комірці кілька дат (розділених пробілами або новими рядками)
                # Розділяємо на окремі дати
                date_strings = re.split(r'[\s\n\r]+', cell_str)
                dates = []
                
                for date_str in date_strings:
                    date_str = date_str.strip()
                    if not date_str:
                        continue
                    # Спробуємо формат dd.mm.yyyy
                    try:
                        date_obj = datetime.strptime(date_str, "%d.%m.%Y")
                        dates.append(date_obj)
                    except ValueError:
                        continue
                
                if dates:
                    # Якщо потрібна перша дата - повертаємо мінімальну
                    # Якщо потрібна остання - повертаємо максимальну
                    if take_first:
                        return min(dates)
                    else:
                        return max(dates)
                
                # Якщо не вдалося розпарсити як окремі дати, спробуємо через parse_date_from_excel_cell
                return parse_date_from_excel_cell(cell_value)
            
            # Для інших типів (наприклад, числа Excel)
            return parse_date_from_excel_cell(cell_value)
        except Exception as e:
            print(f"Помилка парсингу дати {cell_value}: {e}")
            return None
    
    def read_category_sheet(self, sheet_name: str) -> List[Tuple[str, str, str, datetime, datetime]]:
        """
        Читає дані з аркуша категорії (100к, 30к, 0к)
        
        Returns:
            List[Tuple[rank, pib, position, start_date, end_date]]
        """
        if not self.wb:
            self.load_workbook()
        
        if sheet_name not in self.wb.sheetnames:
            print(f"Попередження: аркуш '{sheet_name}' не знайдено")
            return []
        
        ws = self.wb[sheet_name]
        results = []
        
        # Читаємо дані з рядка 2 і далі
        for row in range(2, ws.max_row + 1):
            cell_b = ws.cell(row, 2).value  # Звання + ПІБ + посада
            cell_c = ws.cell(row, 3).value  # Дата початку
            cell_d = ws.cell(row, 4).value  # Дата кінця
            cell_f = ws.cell(row, 6).value if ws.max_column >= 6 else None  # ПІБ окремо
            
            # Перевіряємо, чи є дані
            if not cell_b and not cell_f:
                continue
            
            # Парсимо інформацію про військовослужбовця
            rank, pib, position = self.parse_soldier_info(cell_b, cell_f)
            
            if not pib:
                continue
            
            # Парсимо дати
            # Для початку періоду беремо першу дату, для кінця - останню
            start_date = self.parse_date(cell_c, take_first=True)
            end_date = self.parse_date(cell_d, take_first=False)
            
            if not start_date or not end_date:
                print(f"Попередження: не вдалося розпарсити дати для рядка {row}, ПІБ: {pib}")
                continue
            
            results.append((rank, pib, position, start_date, end_date))
        
        print(f"Прочитано {len(results)} записів з аркуша '{sheet_name}'")
        return results
    
    def read_all_categories(self) -> Dict[str, SoldierPeriod]:
        """
        Читає дані з усіх аркушів категорій та об'єднує по ПІБ
        
        Returns:
            Dict[pib, SoldierPeriod] - словник з даними по кожному ПІБ
        """
        soldiers: Dict[str, SoldierPeriod] = {}
        
        categories = ["100к", "30к", "0к"]
        
        for category in categories:
            records = self.read_category_sheet(category)
            
            for rank, pib, position, start_date, end_date in records:
                # Нормалізуємо ПІБ (великі літери, без зайвих пробілів)
                pib_normalized = " ".join(pib.upper().split())
                
                # Створюємо або отримуємо об'єкт SoldierPeriod
                if pib_normalized not in soldiers:
                    soldiers[pib_normalized] = SoldierPeriod(pib_normalized, rank, position)
                else:
                    # Якщо вже є - оновлюємо звання та посаду, якщо вони були порожні
                    soldier = soldiers[pib_normalized]
                    if not soldier.rank and rank:
                        soldier.rank = rank
                    if not soldier.position and position:
                        soldier.position = position
                
                soldier = soldiers[pib_normalized]
                
                # Додаємо період до відповідної категорії
                period = (start_date, end_date)
                if category == "100к":
                    soldier.periods_100.append(period)
                elif category == "30к":
                    soldier.periods_30.append(period)
                elif category == "0к":
                    soldier.periods_0.append(period)
        
        print(f"Всього унікальних військовослужбовців: {len(soldiers)}")
        return soldiers


class PeriodCollector:
    """Формує позначки за дні місяця на основі періодів"""
    
    def __init__(self, year: int, month: int):
        self.year = year
        self.month = month
        self.days_in_month = self._get_days_in_month(year, month)
    
    def _get_days_in_month(self, year: int, month: int) -> int:
        """Повертає кількість днів у місяці"""
        import calendar
        return calendar.monthrange(year, month)[1]
    
    def get_day_mark(self, soldier: SoldierPeriod, day: int) -> str:
        """
        Повертає позначку для конкретного дня для військовослужбовця
        
        Args:
            soldier: Об'єкт SoldierPeriod
            day: День місяця (1-31)
            
        Returns:
            Позначка: "100", "30", "н/п" або ""
        """
        target_date = datetime(self.year, self.month, day)
        
        # Перевіряємо категорії в пріоритетному порядку: 100, 30, 0
        # Якщо день потрапляє в кілька категорій, використовуємо найвищу
        
        # Перевіряємо 100
        for start, end in soldier.periods_100:
            if start <= target_date <= end:
                return "100"
        
        # Перевіряємо 30
        for start, end in soldier.periods_30:
            if start <= target_date <= end:
                return "30"
        
        # Перевіряємо 0
        for start, end in soldier.periods_0:
            if start <= target_date <= end:
                return "н/п"
        
        return ""
    
    def generate_day_marks(self, soldier: SoldierPeriod) -> List[str]:
        """
        Генерує список позначок для всіх днів місяця
        
        Returns:
            List[str] - список позначок для днів 1-31 (порожні рядки для неіснуючих днів)
        """
        marks = []
        for day in range(1, self.days_in_month + 1):
            mark = self.get_day_mark(soldier, day)
            marks.append(mark)
        
        return marks


class TabelSheetWriter:
    """Клас для запису даних у багатомісячний табель"""
    
    def __init__(self, tabel_file: str):
        self.tabel_file = tabel_file
        self.wb = None
        
    def load_workbook(self):
        """Завантажує Excel файл"""
        self.wb = openpyxl.load_workbook(self.tabel_file)
        print(f"Завантажено файл табелю: {self.tabel_file}")
    
    def clear_sheet_data(self, sheet_name: str, start_row: int = 9):
        """Очищає дані на аркуші від start_row і далі"""
        if sheet_name not in self.wb.sheetnames:
            print(f"Попередження: аркуш '{sheet_name}' не знайдено")
            return
        
        ws = self.wb[sheet_name]
        
        # Очищаємо стовпці D, E, F та G:AK
        for row in range(start_row, ws.max_row + 1):
            # Стовпці D, E, F (посада, звання, ПІБ)
            for col in [4, 5, 6]:
                ws.cell(row, col).value = None
            # Стовпці G:AK (дні місяця, максимально до 31 дня)
            for col in range(7, 38):  # G=7, AK=37 (за 31 день)
                ws.cell(row, col).value = None
    
    def fill_month_sheet(self, sheet_name: str, soldiers: Dict[str, SoldierPeriod], year: int, month: int):
        """
        Заповнює аркуш місяця даними
        
        Args:
            sheet_name: Назва аркуша (наприклад, "Травень_2025")
            soldiers: Словник з даними військовослужбовців
            year: Рік
            month: Місяць
        """
        if not self.wb:
            self.load_workbook()
        
        if sheet_name not in self.wb.sheetnames:
            raise ValueError(f"Аркуш '{sheet_name}' не знайдено у файлі табелю")
        
        ws = self.wb[sheet_name]
        
        # Очищаємо старі дані
        self.clear_sheet_data(sheet_name, start_row=9)
        
        # Створюємо PeriodCollector
        collector = PeriodCollector(year, month)
        
        # Записуємо дані, починаючи з рядка 9
        row = 9
        
        # Сортуємо ПІБ за алфавітом
        sorted_pibs = sorted(soldiers.keys())
        
        for pib in sorted_pibs:
            soldier = soldiers[pib]
            
            # Записуємо дані в рядок
            ws.cell(row, 4).value = soldier.position  # Стовпець D
            ws.cell(row, 5).value = soldier.rank      # Стовпець E
            ws.cell(row, 6).value = soldier.pib       # Стовпець F
            
            # Генеруємо позначки за дні
            day_marks = collector.generate_day_marks(soldier)
            
            # Записуємо позначки в стовпці G:AK
            # G = 7, тому індекс 0 відповідає колонці 7
            for day_idx, mark in enumerate(day_marks):
                col = 7 + day_idx  # Стовпець G = 7
                if mark:  # Записуємо тільки якщо є позначка
                    ws.cell(row, col).value = mark
            
            row += 1
        
        print(f"Заповнено аркуш '{sheet_name}': {len(sorted_pibs)} військовослужбовців")
    
    def save(self):
        """Зберігає файл"""
        if self.wb:
            self.wb.save(self.tabel_file)
            print(f"Файл збережено: {self.tabel_file}")


def fill_tabel_months(tabel_file: str = "Табель_Багатомісячний.xlsx"):
    """
    Основна функція для заповнення аркушів місяців у багатомісячному табелі.
    Автоматично визначає доступні місяці з аркушів Excel.

    Args:
        tabel_file: Шлях до файлу багатомісячного табелю
    """
    from path_utils import get_app_dir
    tabel_path = os.path.join(get_app_dir(), tabel_file) if not os.path.isabs(tabel_file) else tabel_file

    available = get_available_months(tabel_path)
    if not available:
        print("Не знайдено жодного аркуша місяця в табелі")
        return

    writer = TabelSheetWriter(tabel_file)

    for sheet_name in available:
        parsed = parse_month_sheet_name(sheet_name)
        if not parsed:
            continue
        year, month = parsed
        source_file = get_source_filename(sheet_name)
        print(f"\n{'='*60}")
        print(f"Обробка місяця: {sheet_name}")
        print(f"{'='*60}")
        
        try:
            # Читаємо дані з джерела
            reader = SourceFileReader(source_file)
            soldiers = reader.read_all_categories()
            
            # Записуємо в табель
            writer.fill_month_sheet(sheet_name, soldiers, year, month)
            
        except FileNotFoundError as e:
            print(f"Помилка: файл '{source_file}' не знайдено: {e}")
        except Exception as e:
            print(f"Помилка при обробці {sheet_name}: {e}")
            import traceback
            traceback.print_exc()
    
    # Зберігаємо файл
    writer.save()
    print(f"\n{'='*60}")
    print("Заповнення завершено!")
    print(f"{'='*60}")


def fill_single_month(sheet_name: str, source_file: str, year: int, month: int, 
                     tabel_file: str = "Табель_Багатомісячний.xlsx"):
    """
    Заповнює один аркуш місяця
    
    Args:
        sheet_name: Назва аркуша в табелі
        source_file: Шлях до джерельного файлу
        year: Рік
        month: Місяць
        tabel_file: Шлях до файлу багатомісячного табелю
    """
    print(f"\n{'='*60}")
    print(f"Обробка місяця: {sheet_name}")
    print(f"{'='*60}")
    
    try:
        # Читаємо дані з джерела
        reader = SourceFileReader(source_file)
        soldiers = reader.read_all_categories()
        
        # Записуємо в табель
        writer = TabelSheetWriter(tabel_file)
        writer.fill_month_sheet(sheet_name, soldiers, year, month)
        writer.save()
        
        print(f"\n{'='*60}")
        print(f"Заповнення {sheet_name} завершено!")
        print(f"{'='*60}")
        
    except Exception as e:
        print(f"Помилка: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    import sys

    tabel_file = "Табель_Багатомісячний.xlsx"
    from path_utils import get_app_dir
    tabel_path = os.path.join(get_app_dir(), tabel_file)

    if len(sys.argv) > 1:
        # Якщо передано аргумент - заповнюємо конкретний місяць
        month_arg = sys.argv[1].lower()
        available = get_available_months(tabel_path)

        matched = None
        for sheet_name in available:
            parsed = parse_month_sheet_name(sheet_name)
            if parsed and MONTH_NAMES_UK_LOWER[parsed[1]] == month_arg:
                year, month_num = parsed
                matched = (sheet_name, get_source_filename(sheet_name), year, month_num)
                break

        if matched:
            fill_single_month(*matched)
        else:
            available_names = [MONTH_NAMES_UK_LOWER[parse_month_sheet_name(s)[1]] for s in available if parse_month_sheet_name(s)]
            print(f"Невідомий місяць: {month_arg}")
            print(f"Доступні місяці: {', '.join(available_names)}")
    else:
        # За замовчуванням - заповнюємо всі місяці
        fill_tabel_months()

