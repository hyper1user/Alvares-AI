"""
Модуль читання даних з місячних табелів
"""
from month_utils import MONTH_NAMES_UK, parse_month_sheet_name
import openpyxl
from datetime import datetime
from typing import List, Dict, Any, Optional
from br_calculator import parse_date_from_excel_cell, get_day_column_for_date, get_br_number, format_br_list

class SoldierData:
    """Клас для зберігання даних про бійця"""
    
    def __init__(self, row_number: int, pib: str, rank: str, position: str):
        self.row_number = row_number
        self.pib = pib
        self.rank = rank
        self.position = position
        self.days_100: List[datetime] = []  # Дні з позначкою 100
        self.days_rop: List[datetime] = []  # Дні з позначкою роп (позиції, прирівнюється до 100)
        self.days_30: List[datetime] = []   # Дні з позначкою 30
        self.days_0: List[datetime] = []    # Дні з позначкою н-п
        self.note: str = ""                 # Примітка (може містити "не виплачувати")
        self.br_numbers_100: List[str] = [] # Номери БР для днів 100 (включаючи роп)
        self.br_numbers_30: List[str] = []  # Номери БР для днів 30

    def add_day(self, date: datetime, mark: str):
        """Додає день з позначкою"""
        if mark == "100":
            self.days_100.append(date)
        elif mark == "роп":
            self.days_rop.append(date)
        elif mark == "30":
            self.days_30.append(date)
        elif mark in ["0", "н-п", "н/п"]:
            self.days_0.append(date)

    @property
    def days_100_combined(self) -> List[datetime]:
        """Всі дні 100 + роп (для табелю/виплат)"""
        return sorted(self.days_100 + self.days_rop)

    def generate_br_numbers(self):
        """Генерує номери БР для всіх днів"""
        self.br_numbers_100 = [get_br_number(date) for date in self.days_100_combined]
        self.br_numbers_30 = [get_br_number(date) for date in self.days_30]
    
    def get_br_list_100(self) -> str:
        """Повертає відформатований список БР для днів 100"""
        return format_br_list(self.br_numbers_100)
    
    def get_br_list_30(self) -> str:
        """Повертає відформатований список БР для днів 30"""
        return format_br_list(self.br_numbers_30)
    
    def has_no_payment_note(self) -> bool:
        """Перевіряє чи є примітка про невиплату"""
        return "не виплачувати" in self.note.lower() if self.note else False
    
    def __repr__(self):
        return f"SoldierData({self.pib}, 100:{len(self.days_100)}, 30:{len(self.days_30)}, 0:{len(self.days_0)})"

class TabelReader:
    """Клас для читання даних з табелю"""
    
    def __init__(self, excel_file: str):
        self.excel_file = excel_file
        self.wb = None
        self.soldiers: List[SoldierData] = []
    
    def load_workbook(self):
        """Завантажує Excel файл"""
        self.wb = openpyxl.load_workbook(self.excel_file, data_only=True)
        print(f"Завантажено файл: {self.excel_file}")
        print(f"Листи: {self.wb.sheetnames}")
    
    def read_month_data(self, month_sheet: str) -> List[SoldierData]:
        """
        Читає дані з листа конкретного місяця
        
        Args:
            month_sheet: Назва листа (наприклад, "Вересень_2025")
            
        Returns:
            List[SoldierData]: Список даних про бійців
        """
        if not self.wb:
            self.load_workbook()
        
        if month_sheet not in self.wb.sheetnames:
            raise ValueError(f"Лист '{month_sheet}' не знайдено")
        
        ws = self.wb[month_sheet]
        soldiers = []
        
        # Знаходимо рядок з заголовками (шукаємо "ПІБ")
        header_row = None
        for row in range(1, 20):
            if ws.cell(row, 6).value and "ПІБ" in str(ws.cell(row, 6).value):
                header_row = row
                break
        
        if not header_row:
            raise ValueError("Не знайдено рядок з заголовками")
        
        print(f"Заголовки знайдено в рядку {header_row}")

        # Визначаємо рік та місяць з назви листа
        year = self._extract_year_from_sheet_name(month_sheet)
        month = self._extract_month_from_sheet_name(month_sheet)
        
        # Читаємо дані бійців
        for row in range(header_row + 1, ws.max_row + 1):
            pib = ws.cell(row, 6).value  # Стовпець F
            if not pib or not str(pib).strip():
                continue
            
            rank = ws.cell(row, 5).value or ""  # Стовпець E
            position = ws.cell(row, 4).value or ""  # Стовпець D
            note = ws.cell(row, 38).value or ""  # Стовпець AL
            
            soldier = SoldierData(row, str(pib), str(rank), str(position))
            soldier.note = str(note)
            
            # Читаємо позначки за дні
            days_in_month = self._get_days_in_month(year, month)
            for day in range(1, days_in_month + 1):
                col = 6 + day  # Стовпець G = 7, тому 6 + day
                cell_value = ws.cell(row, col).value
                
                if cell_value:
                    mark = str(cell_value).strip().lower()
                    # Нормалізуємо: "100" і "30" зберігаємо як є, "роп" окремо
                    if mark in ["100", "30", "0", "н-п", "н/п", "роп"]:
                        try:
                            date = datetime(year, month, day)
                            soldier.add_day(date, mark)
                        except ValueError:
                            print(f"Помилка дати для рядка {row}, день {day}")
            
            soldier.generate_br_numbers()
            soldiers.append(soldier)
        
        print(f"Прочитано {len(soldiers)} бійців з листа {month_sheet}")
        return soldiers
    
    def _extract_month_from_sheet_name(self, sheet_name: str) -> int:
        """Витягує номер місяця з назви листа"""
        parsed = parse_month_sheet_name(sheet_name)
        if parsed:
            return parsed[1]
        # Fallback: пошук за підрядком
        sheet_lower = sheet_name.lower()
        for name, num in MONTH_NAMES_UK.items():
            if name in sheet_lower:
                return num
        raise ValueError(f"Не можу визначити місяць з назви листа: {sheet_name}")

    def _extract_year_from_sheet_name(self, sheet_name: str) -> int:
        """Витягує рік з назви листа (наприклад, 'Січень_2026' -> 2026)"""
        parsed = parse_month_sheet_name(sheet_name)
        if parsed:
            return parsed[0]
        import re
        match = re.search(r'_(\d{4})', sheet_name)
        if match:
            return int(match.group(1))
        return 2025

    def _get_days_in_month(self, year: int, month: int) -> int:
        """Повертає кількість днів у місяці"""
        import calendar
        return calendar.monthrange(year, month)[1]
    
    def get_soldiers_by_category(self, soldiers: List[SoldierData], category: str, include_no_payment: bool = False) -> List[SoldierData]:
        """
        Фільтрує бійців за категорією
        
        Args:
            soldiers: Список бійців
            category: "100", "30", "0"
            include_no_payment: Чи включати бійців з приміткою "не виплачувати"
            
        Returns:
            List[SoldierData]: Відфільтрований список
        """
        result = []
        
        for soldier in soldiers:
            if category == "100" and soldier.days_100_combined:
                if include_no_payment or not soldier.has_no_payment_note():
                    result.append(soldier)
            elif category == "30" and soldier.days_30:
                if include_no_payment or not soldier.has_no_payment_note():
                    result.append(soldier)
            elif category == "0" and soldier.days_0:
                if include_no_payment or not soldier.has_no_payment_note():
                    result.append(soldier)
        
        return result
    
    def get_period_string(self, dates: List[datetime]) -> str:
        """Повертає рядок періоду для рапорту"""
        if not dates:
            return ""
        
        dates.sort()
        start_date = dates[0]
        end_date = dates[-1]
        
        if start_date == end_date:
            return start_date.strftime("%d.%m.%Y")
        else:
            return f"з {start_date.strftime('%d.%m.%Y')} по {end_date.strftime('%d.%m.%Y')}"

# Приклад використання
if __name__ == "__main__":
    # Тестування
    reader = TabelReader("Табель_Багатомісячний.xlsx")
    
    try:
        soldiers = reader.read_month_data("Травень_2025")
        
        print(f"\nВсього бійців: {len(soldiers)}")
        
        # Тест фільтрації
        soldiers_100 = reader.get_soldiers_by_category(soldiers, "100", include_no_payment=False)
        soldiers_30 = reader.get_soldiers_by_category(soldiers, "30", include_no_payment=False)
        
        print(f"Військовслужбовців 12ШР 4 ШБ на 100 (без 'не виплачувати'): {len(soldiers_100)}")
        print(f"Військовслужбовців 12ШР 4 ШБ на 30 (без 'не виплачувати'): {len(soldiers_30)}")
        
        # Показуємо першого бійця як приклад
        if soldiers:
            soldier = soldiers[0]
            print(f"\nПриклад бійця: {soldier.pib}")
            print(f"  Дні 100: {len(soldier.days_100)}")
            print(f"  Дні 30: {len(soldier.days_30)}")
            print(f"  Дні 0: {len(soldier.days_0)}")
            if soldier.br_numbers_100:
                print(f"  БР для 100: {soldier.get_br_list_100()[:100]}...")
    
    except Exception as e:
        print(f"Помилка: {e}")

