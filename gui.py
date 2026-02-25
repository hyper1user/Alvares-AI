"""
Графічний інтерфейс для системи генерації рапортів ДГВ
"""
import customtkinter as ctk
import tkinter as tk
from tkinter import ttk, messagebox
import threading
import os
import openpyxl
from datetime import datetime
from typing import Optional
import webbrowser

try:
    from PIL import Image, ImageTk
    PIL_AVAILABLE = True
except ImportError:
    PIL_AVAILABLE = False

from generate_reports import ReportGenerator
from br_calculator import get_br_number
from month_utils import (get_available_months, parse_month_sheet_name, get_source_filename,
                         build_month_sheet_name, MONTH_NAMES_UK_REVERSE)
from tabel_filler import fill_single_month, fill_tabel_months
from data.database import (init_db, get_all_personnel, get_all_roles,
                           set_personnel_role)
from core.br_roles import (auto_assign_all_roles, import_personnel_from_tabel,
                           build_composition_for_date, generate_br_word,
                           generate_rop_word, get_active_personnel_for_month)
from path_utils import get_base_path, get_app_dir
from version import APP_VERSION
from updater import check_for_update, get_releases_url, download_update, install_update

# Налаштування теми
ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")

# Акцентні кольори для кнопок
_CLR_BLUE = "#3498db"
_CLR_BLUE_HOVER = "#2980b9"
_CLR_ORANGE = "#e67e22"
_CLR_ORANGE_HOVER = "#d35400"
_CLR_RED = "#c0392b"
_CLR_RED_HOVER = "#a93226"
_CLR_PURPLE = "#8e44ad"
_CLR_PURPLE_HOVER = "#7d3c98"
_CLR_TEAL = "#16a085"
_CLR_TEAL_HOVER = "#138d75"
_CLR_GREEN = "#27ae60"
_CLR_GREEN_HOVER = "#229954"
_CLR_GRAY = "#95a5a6"
_CLR_GRAY_HOVER = "#7f8c8d"
_CLR_YELLOW = "#f1c40f"
_CLR_DIM = "#7f849c"

# Стилі для Treeview (залишаємо ttk)
_TREE_BG = "#2b2b2b"
_TREE_FG = "#dce4ee"
_TREE_SELECTED = "#45475a"
_TREE_HEADING_BG = "#1a1a2e"


class ReportGUI:
    """Графічний інтерфейс для генерації рапортів"""

    def __init__(self, root: ctk.CTk):
        self.root = root
        self.root.title("АЛЬВАРЕС AI — Система обліку особового складу")
        self.root.geometry("850x750")
        self.root.resizable(True, True)

        # Іконка вікна
        ico_path = os.path.join(get_base_path(), "alvares.ico")
        if os.path.exists(ico_path):
            self.root.iconbitmap(ico_path)

        # Стилізація Treeview (ttk) для екрану ролей
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Dark.Treeview",
                         background=_TREE_BG, foreground=_TREE_FG,
                         fieldbackground=_TREE_BG, rowheight=28,
                         font=("Arial", 10))
        style.configure("Dark.Treeview.Heading",
                         background=_TREE_HEADING_BG, foreground=_TREE_FG,
                         font=("Arial", 10, "bold"))
        style.map("Dark.Treeview",
                   background=[("selected", _TREE_SELECTED)],
                   foreground=[("selected", "#cdd6f4")])

        # Шляхи
        base_path = get_base_path()
        app_dir = get_app_dir()

        self.generator = None
        self.excel_file = os.path.join(app_dir, "Табель_Багатомісячний.xlsx")

        # Змінні
        self.selected_month = tk.StringVar()
        self.report_type_var = tk.StringVar(value="6")
        self.start_date_var = tk.StringVar()
        self.end_date_var = tk.StringVar()
        self.order_number_var = tk.StringVar()
        self.logo_image = None
        self._busy_photo = None

        self.current_screen = None

        # БД ролей
        init_db()

        self.template_path = os.path.join(base_path, "templates", "rozp_template.docx")
        self.rop_template_path = os.path.join(base_path, "templates", "pozition_template.docx")
        self.br_4shb_file = os.path.join(app_dir, "BR_4ShB.xlsx")
        self.output_dir = os.path.join(app_dir, "output")

        self._create_main_menu()
        self._check_files()

        # Перевірка оновлень
        threading.Thread(target=self._background_update_check, daemon=True).start()

    # ==================== ПЕРЕВІРКА ОНОВЛЕНЬ ====================

    def _background_update_check(self):
        update = check_for_update()
        if update:
            self.root.after(0, lambda: self._show_update_notification(update))

    def _manual_update_check(self):
        if hasattr(self, "_version_label"):
            self._version_label.configure(text="Перевірка...", text_color=_CLR_DIM)

        def do_check():
            update = check_for_update()
            if update:
                self.root.after(0, lambda: self._show_update_notification(update))
            else:
                def restore():
                    if hasattr(self, "_version_label"):
                        self._version_label.configure(text=f"v{APP_VERSION}  ✓", text_color="#2ecc71")
                        self.root.after(3000, lambda: self._version_label.configure(
                            text=f"v{APP_VERSION}", text_color=_CLR_DIM
                        ) if hasattr(self, "_version_label") else None)
                self.root.after(0, restore)

        threading.Thread(target=do_check, daemon=True).start()

    def _show_update_notification(self, update: dict):
        version = update["version"]
        notes = update.get("notes", "")

        if hasattr(self, "_version_label"):
            self._version_label.configure(
                text=f"v{APP_VERSION}  →  v{version} доступна!",
                text_color=_CLR_YELLOW
            )

        notes_snippet = (notes[:300] + "...") if len(notes) > 300 else notes
        msg = f"Доступна нова версія АЛЬВАРЕС AI!\n\nПоточна: v{APP_VERSION}\nНова:     v{version}"
        if notes_snippet:
            msg += f"\n\nЩо нового:\n{notes_snippet}"

        if update.get("download_url"):
            msg += "\n\nОновити автоматично?"
            if messagebox.askyesno("Оновлення доступне", msg):
                self._do_update(update)
        else:
            msg += "\n\nВідкрити сторінку завантаження?"
            if messagebox.askyesno("Оновлення доступне", msg):
                webbrowser.open(update["url"])

    def _do_update(self, update: dict):
        download_url = update["download_url"]
        version = update["version"]

        win = ctk.CTkToplevel(self.root)
        win.title(f"Оновлення до v{version}")
        win.geometry("420x150")
        win.resizable(False, False)
        win.transient(self.root)
        win.grab_set()

        status_label = ctk.CTkLabel(win, text="Завантаження...", font=ctk.CTkFont(size=14))
        status_label.pack(pady=(20, 5))

        progress_bar = ctk.CTkProgressBar(win, width=360)
        progress_bar.pack(pady=5)
        progress_bar.set(0)

        size_label = ctk.CTkLabel(win, text="0 / ? МБ", font=ctk.CTkFont(size=12), text_color=_CLR_DIM)
        size_label.pack(pady=5)

        def on_progress(downloaded, total):
            dl_mb = downloaded / (1024 * 1024)
            if total > 0:
                total_mb = total / (1024 * 1024)
                fraction = downloaded / total
                self.root.after(0, lambda: progress_bar.set(fraction))
                self.root.after(0, lambda: size_label.configure(text=f"{dl_mb:.1f} / {total_mb:.1f} МБ"))
            else:
                self.root.after(0, lambda: size_label.configure(text=f"{dl_mb:.1f} МБ"))

        def do_download():
            try:
                setup_path = download_update(download_url, on_progress=on_progress)
                self.root.after(0, lambda: _on_download_complete(setup_path))
            except Exception as e:
                self.root.after(0, lambda: _on_download_error(str(e)))

        def _on_download_complete(setup_path):
            status_label.configure(text="Встановлення оновлення...")
            progress_bar.set(1.0)
            win.after(500, lambda: install_update(setup_path))

        def _on_download_error(err):
            status_label.configure(text="Помилка завантаження")
            size_label.configure(text=err[:80], text_color="#e74c3c")
            win.after(3000, win.destroy)

        threading.Thread(target=do_download, daemon=True).start()

    # ==================== УТИЛІТИ ====================

    def _get_logo(self):
        if not PIL_AVAILABLE:
            return None
        if self.logo_image:
            return self.logo_image
        emblem_path = os.path.join(get_base_path(), "emblem.png")
        if not os.path.exists(emblem_path):
            return None
        try:
            img = Image.open(emblem_path)
            target_h = 44
            w, h = img.size
            target_w = max(1, int(w * (target_h / float(h))))
            img = img.resize((target_w, target_h), Image.Resampling.LANCZOS)
            self.logo_image = ImageTk.PhotoImage(img)
            return self.logo_image
        except Exception:
            return None

    def _clear_screen(self):
        for widget in self.root.winfo_children():
            widget.destroy()

    def _make_header(self, parent, title_text: str) -> ctk.CTkFrame:
        """Створює стандартний заголовок з логотипом та назвою."""
        header = ctk.CTkFrame(parent, height=70, corner_radius=0,
                               fg_color=("#e0e0e0", "#1a1a2e"))
        header.pack(fill="x", pady=(0, 15))
        header.pack_propagate(False)

        logo = self._get_logo()
        if logo:
            logo_lbl = tk.Label(header, image=logo, bg="#1a1a2e")
            logo_lbl.image = logo
            logo_lbl.pack(side="left", padx=12)

        ctk.CTkLabel(
            header, text=title_text,
            font=ctk.CTkFont(size=20, weight="bold")
        ).pack(pady=18)

        return header

    # ==================== ГОЛОВНЕ МЕНЮ ====================

    def _create_main_menu(self):
        self.current_screen = "main"
        self._clear_screen()

        content = ctk.CTkFrame(self.root, fg_color="transparent")
        content.pack(fill="both", expand=True)

        # Заголовок
        header = ctk.CTkFrame(content, height=80, corner_radius=0,
                               fg_color=("#e0e0e0", "#1a1a2e"))
        header.pack(fill="x", pady=(0, 25))
        header.pack_propagate(False)

        logo = self._get_logo()
        if logo:
            logo_lbl = tk.Label(header, image=logo, bg="#1a1a2e")
            logo_lbl.image = logo
            logo_lbl.pack(side="left", padx=12)

        ctk.CTkLabel(
            header, text="АЛЬВАРЕС AI",
            font=ctk.CTkFont(size=22, weight="bold")
        ).pack(pady=(15, 2))

        ctk.CTkLabel(
            header, text="Система обліку особового складу 12 штурмової роти",
            font=ctk.CTkFont(size=12), text_color=_CLR_DIM
        ).pack()

        # Мітка версії — клікабельна
        self._version_label = ctk.CTkLabel(
            header, text=f"v{APP_VERSION}",
            font=ctk.CTkFont(size=10), text_color=_CLR_DIM, cursor="hand2"
        )
        self._version_label.place(relx=1.0, rely=0.0, anchor="ne", x=-10, y=6)
        self._version_label.bind("<Button-1>", lambda e: self._manual_update_check())
        self._version_label.bind("<Enter>", lambda e: self._version_label.configure(text_color="#ffffff"))
        self._version_label.bind("<Leave>", lambda e: self._version_label.configure(text_color=_CLR_DIM))

        # Кнопки
        buttons_frame = ctk.CTkFrame(content, fg_color="transparent")
        buttons_frame.pack(fill="both", expand=True, padx=60, pady=10)

        menu_buttons = [
            ("📄  Створити документи за місяць", _CLR_BLUE, _CLR_BLUE_HOVER, self._show_reports_screen),
            ("📋  Створити БР", _CLR_ORANGE, _CLR_ORANGE_HOVER, self._show_br_create_screen),
            ("🧩  Ролі для БР", _CLR_RED, _CLR_RED_HOVER, self._show_roles_screen),
            ("📊  Заповнити табель", _CLR_PURPLE, _CLR_PURPLE_HOVER, self._show_tabel_filler_screen),
            ("➕  Додати місяць", _CLR_TEAL, _CLR_TEAL_HOVER, self._show_add_month_dialog),
        ]

        for text, color, hover, cmd in menu_buttons:
            ctk.CTkButton(
                buttons_frame, text=text, command=cmd,
                font=ctk.CTkFont(size=15, weight="bold"),
                fg_color=color, hover_color=hover,
                height=55, corner_radius=12
            ).pack(fill="x", pady=8)

        # Статус-бар
        self._make_status_bar(content, "Готово до роботи")

    # ==================== ЕКРАН ДОКУМЕНТІВ ====================

    def _show_reports_screen(self):
        self.current_screen = "reports"
        self._clear_screen()

        content = ctk.CTkFrame(self.root, fg_color="transparent")
        content.pack(fill="both", expand=True)

        self._make_header(content, "Створити документи за місяць")

        main = ctk.CTkFrame(content, fg_color="transparent")
        main.pack(fill="both", expand=True, padx=30, pady=(0, 10))

        # Вибір місяця
        ctk.CTkLabel(main, text="Оберіть місяць", font=ctk.CTkFont(size=13, weight="bold")).pack(anchor="w", pady=(0, 5))
        months = self.generator.available_months if self.generator else get_available_months(self.excel_file)
        month_combo = ctk.CTkComboBox(main, variable=self.selected_month, values=months,
                                       width=350, font=ctk.CTkFont(size=12), state="readonly")
        month_combo.pack(pady=(0, 15))
        if months:
            month_combo.set(months[0])

        # Тип рапорту
        ctk.CTkLabel(main, text="Оберіть тип рапорту", font=ctk.CTkFont(size=13, weight="bold")).pack(anchor="w", pady=(0, 5))

        report_types = {
            "1": "ДГВ 100к (Excel)",
            "2": "ДГВ 30к (Excel)",
            "3": "Підтвердження 100к (Word)",
            "4": "Підтвердження 30к (Word)",
            "5": "ДГВ 0к (Excel)",
            "6": "Створити всі типи за обраний місяць"
        }

        radio_frame = ctk.CTkFrame(main, fg_color="transparent")
        radio_frame.pack(fill="x", pady=(0, 15))
        for key, desc in report_types.items():
            ctk.CTkRadioButton(
                radio_frame, text=desc, variable=self.report_type_var, value=key,
                font=ctk.CTkFont(size=12)
            ).pack(anchor="w", pady=3)

        # Кнопки
        btn_frame = ctk.CTkFrame(main, fg_color="transparent")
        btn_frame.pack(fill="x", pady=(0, 10))

        self.generate_btn = ctk.CTkButton(
            btn_frame, text="📄 Створити рапорти", command=self._generate_reports,
            font=ctk.CTkFont(size=13, weight="bold"),
            fg_color=_CLR_GREEN, hover_color=_CLR_GREEN_HOVER, height=42
        )
        self.generate_btn.pack(side="left", padx=(0, 10))

        ctk.CTkButton(
            btn_frame, text="← Назад", command=self._create_main_menu,
            font=ctk.CTkFont(size=12),
            fg_color=_CLR_GRAY, hover_color=_CLR_GRAY_HOVER, height=42
        ).pack(side="right")

        # Лог
        ctk.CTkLabel(main, text="Статус виконання", font=ctk.CTkFont(size=13, weight="bold")).pack(anchor="w", pady=(0, 5))
        self.log_text = ctk.CTkTextbox(main, font=ctk.CTkFont(family="Consolas", size=10), height=150)
        self.log_text.pack(fill="both", expand=True)

        self._make_status_bar(content, "Готово до роботи")

    # ==================== ЕКРАН БР ====================

    def _show_br_create_screen(self):
        self.current_screen = "br_create"
        self._clear_screen()

        content = ctk.CTkFrame(self.root, fg_color="transparent")
        content.pack(fill="both", expand=True)

        self._make_header(content, "Створити БР")

        main = ctk.CTkFrame(content, fg_color="transparent")
        main.pack(fill="both", expand=True, padx=30, pady=(0, 10))

        # Дати
        ctk.CTkLabel(main, text="Оберіть період", font=ctk.CTkFont(size=13, weight="bold")).pack(anchor="w", pady=(0, 5))

        dates_frame = ctk.CTkFrame(main, fg_color="transparent")
        dates_frame.pack(fill="x", pady=(0, 15))

        # Початкова дата
        row1 = ctk.CTkFrame(dates_frame, fg_color="transparent")
        row1.pack(fill="x", pady=3)
        ctk.CTkLabel(row1, text="Початкова дата (ДД.ММ.РРРР):", font=ctk.CTkFont(size=11)).pack(side="left", padx=(0, 10))
        start_entry = ctk.CTkEntry(row1, textvariable=self.start_date_var, width=180, font=ctk.CTkFont(size=12))
        start_entry.pack(side="left")
        start_entry.bind('<KeyRelease>', lambda e: self._update_order_number())

        # Кінцева дата
        row2 = ctk.CTkFrame(dates_frame, fg_color="transparent")
        row2.pack(fill="x", pady=3)
        ctk.CTkLabel(row2, text="Кінцева дата (ДД.ММ.РРРР):", font=ctk.CTkFont(size=11)).pack(side="left", padx=(0, 10))
        ctk.CTkEntry(row2, textvariable=self.end_date_var, width=180, font=ctk.CTkFont(size=12)).pack(side="left")

        # Номер наказу
        ctk.CTkLabel(main, text="Початковий номер наказу", font=ctk.CTkFont(size=13, weight="bold")).pack(anchor="w", pady=(0, 5))

        order_frame = ctk.CTkFrame(main)
        order_frame.pack(fill="x", pady=(0, 5))
        ctk.CTkLabel(order_frame, textvariable=self.order_number_var,
                      font=ctk.CTkFont(size=13, weight="bold")).pack(padx=15, pady=8)

        ctk.CTkLabel(main, text="(Номер розраховується автоматично)",
                      font=ctk.CTkFont(size=10), text_color=_CLR_DIM).pack(anchor="w", pady=(0, 10))

        # Кнопки
        btn_frame = ctk.CTkFrame(main, fg_color="transparent")
        btn_frame.pack(fill="x", pady=(0, 10))

        ctk.CTkButton(
            btn_frame, text="Сформувати склад", command=self._preview_composition,
            font=ctk.CTkFont(size=12, weight="bold"),
            fg_color=_CLR_BLUE, hover_color=_CLR_BLUE_HOVER, height=40
        ).pack(side="left", padx=(0, 8))

        ctk.CTkButton(
            btn_frame, text="Згенерувати Word БР", command=self._generate_word_br,
            font=ctk.CTkFont(size=12, weight="bold"),
            fg_color=_CLR_GREEN, hover_color=_CLR_GREEN_HOVER, height=40
        ).pack(side="left", padx=(0, 8))

        ctk.CTkButton(
            btn_frame, text="← Назад", command=self._create_main_menu,
            font=ctk.CTkFont(size=12),
            fg_color=_CLR_GRAY, hover_color=_CLR_GRAY_HOVER, height=40
        ).pack(side="right")

        # Лог
        ctk.CTkLabel(main, text="Статус виконання", font=ctk.CTkFont(size=13, weight="bold")).pack(anchor="w", pady=(0, 5))
        self.log_text = ctk.CTkTextbox(main, font=ctk.CTkFont(family="Consolas", size=10), height=150)
        self.log_text.pack(fill="both", expand=True)

        self._make_status_bar(content, "Введіть дати для автоматичного розрахунку номера наказу")

        # Значення за замовчуванням
        today = datetime.now()
        self.start_date_var.set(today.strftime("%d.%m.%Y"))
        self.end_date_var.set(today.strftime("%d.%m.%Y"))
        self._update_order_number()

    def _update_order_number(self):
        date_str = self.start_date_var.get().strip()
        if not date_str:
            self.order_number_var.set("Введіть початкову дату")
            return
        try:
            date_obj = datetime.strptime(date_str, "%d.%m.%Y")
            from br_updater import get_tabel_date
            tabel_date = get_tabel_date(date_obj)
            self.order_number_var.set(get_br_number(tabel_date))
        except ValueError:
            self.order_number_var.set("Некоректний формат дати (ДД.ММ.РРРР)")

    # ==================== ЕКРАН ЗАПОВНЕННЯ ТАБЕЛЮ ====================

    def _show_tabel_filler_screen(self):
        self.current_screen = "tabel_filler"
        self._clear_screen()

        content = ctk.CTkFrame(self.root, fg_color="transparent")
        content.pack(fill="both", expand=True)

        self._make_header(content, "Заповнити табель з місячних файлів")

        main = ctk.CTkFrame(content, fg_color="transparent")
        main.pack(fill="both", expand=True, padx=30, pady=(0, 10))

        # Вибір місяця
        ctk.CTkLabel(main, text="Оберіть місяць", font=ctk.CTkFont(size=13, weight="bold")).pack(anchor="w", pady=(0, 5))
        months = self.generator.available_months if self.generator else get_available_months(self.excel_file)
        month_values = months + ["-- Всі місяці --"]

        self.tabel_month_var = tk.StringVar()
        tabel_combo = ctk.CTkComboBox(main, variable=self.tabel_month_var, values=month_values,
                                       width=350, font=ctk.CTkFont(size=12), state="readonly",
                                       command=lambda _: self._update_source_files_status())
        tabel_combo.pack(pady=(0, 15))
        if months:
            tabel_combo.set(months[-1])

        # Статус файлів
        ctk.CTkLabel(main, text="Файли-джерела", font=ctk.CTkFont(size=13, weight="bold")).pack(anchor="w", pady=(0, 5))
        self.files_status_text = ctk.CTkTextbox(main, height=60, font=ctk.CTkFont(family="Consolas", size=11))
        self.files_status_text.pack(fill="x", pady=(0, 15))
        self._update_source_files_status()

        # Кнопки
        btn_frame = ctk.CTkFrame(main, fg_color="transparent")
        btn_frame.pack(fill="x", pady=(0, 10))

        self.fill_tabel_btn = ctk.CTkButton(
            btn_frame, text="📊 Заповнити табель", command=self._fill_tabel,
            font=ctk.CTkFont(size=13, weight="bold"),
            fg_color=_CLR_PURPLE, hover_color=_CLR_PURPLE_HOVER, height=42
        )
        self.fill_tabel_btn.pack(side="left", padx=(0, 10))

        ctk.CTkButton(
            btn_frame, text="← Назад", command=self._create_main_menu,
            font=ctk.CTkFont(size=12),
            fg_color=_CLR_GRAY, hover_color=_CLR_GRAY_HOVER, height=42
        ).pack(side="right")

        # Лог
        ctk.CTkLabel(main, text="Статус виконання", font=ctk.CTkFont(size=13, weight="bold")).pack(anchor="w", pady=(0, 5))
        self.log_text = ctk.CTkTextbox(main, font=ctk.CTkFont(family="Consolas", size=10), height=150)
        self.log_text.pack(fill="both", expand=True)

        self._make_status_bar(content, "Готово до роботи")

    def _update_source_files_status(self):
        if not hasattr(self, 'files_status_text') or not hasattr(self, 'tabel_month_var'):
            return

        self.files_status_text.configure(state="normal")
        self.files_status_text.delete("0.0", "end")

        selected = self.tabel_month_var.get()
        if selected == "-- Всі місяці --":
            self.files_status_text.insert("0.0", "Буде оброблено всі доступні місяці")
            self.files_status_text.configure(state="disabled")
            return

        if not selected:
            self.files_status_text.configure(state="disabled")
            return

        source_file = get_source_filename(selected)
        source_path = os.path.join(get_app_dir(), source_file)
        exists = os.path.exists(source_path)
        status = "[+] знайдено" if exists else "[-] не знайдено"
        self.files_status_text.insert("0.0", f"{source_file}  {status}")
        self.files_status_text.configure(state="disabled")

    # ==================== ДІАЛОГ ДОДАВАННЯ МІСЯЦЯ ====================

    def _show_add_month_dialog(self):
        dialog = ctk.CTkToplevel(self.root)
        dialog.title("Додати новий місяць")
        dialog.geometry("420x300")
        dialog.resizable(False, False)
        dialog.transient(self.root)
        dialog.grab_set()

        # Заголовок
        ctk.CTkLabel(dialog, text="Додати новий місяць",
                      font=ctk.CTkFont(size=16, weight="bold")).pack(pady=(20, 15))

        form = ctk.CTkFrame(dialog, fg_color="transparent")
        form.pack(fill="both", expand=True, padx=30)

        # Рік
        row_year = ctk.CTkFrame(form, fg_color="transparent")
        row_year.pack(fill="x", pady=8)
        ctk.CTkLabel(row_year, text="Рік:", font=ctk.CTkFont(size=12), width=80, anchor="w").pack(side="left")
        year_values = [str(y) for y in range(2025, 2031)]
        year_var = tk.StringVar(value=str(datetime.now().year))
        ctk.CTkComboBox(row_year, variable=year_var, values=year_values,
                         width=120, state="readonly").pack(side="left", padx=10)

        # Місяць
        row_month = ctk.CTkFrame(form, fg_color="transparent")
        row_month.pack(fill="x", pady=8)
        ctk.CTkLabel(row_month, text="Місяць:", font=ctk.CTkFont(size=12), width=80, anchor="w").pack(side="left")
        month_names_list = [MONTH_NAMES_UK_REVERSE[i] for i in range(1, 13)]
        month_var = tk.StringVar()
        month_combo = ctk.CTkComboBox(row_month, variable=month_var, values=month_names_list,
                                       width=180, state="readonly")
        month_combo.pack(side="left", padx=10)
        current_month = datetime.now().month
        month_combo.set(MONTH_NAMES_UK_REVERSE[current_month])

        def do_create():
            if not month_var.get():
                messagebox.showwarning("Увага", "Оберіть місяць!", parent=dialog)
                return

            year = int(year_var.get())
            month_name = month_var.get()
            sheet_name = f"{month_name}_{year}"

            try:
                wb = openpyxl.load_workbook(self.excel_file)
                if sheet_name in wb.sheetnames:
                    messagebox.showwarning("Увага", f"Аркуш '{sheet_name}' вже існує!", parent=dialog)
                    wb.close()
                    return

                available = [s for s in wb.sheetnames if parse_month_sheet_name(s)]
                if available:
                    template_sheet = wb[available[-1]]
                    new_sheet = wb.copy_worksheet(template_sheet)
                    new_sheet.title = sheet_name
                    for row in range(9, new_sheet.max_row + 1):
                        for col in range(1, new_sheet.max_column + 1):
                            new_sheet.cell(row, col).value = None
                else:
                    wb.create_sheet(sheet_name)

                wb.save(self.excel_file)
                wb.close()

                messagebox.showinfo("Успіх", f"Аркуш '{sheet_name}' створено!", parent=dialog)

                try:
                    self.generator = ReportGenerator(self.excel_file)
                except Exception:
                    pass

                dialog.destroy()

            except Exception as e:
                messagebox.showerror("Помилка", f"Не вдалося створити аркуш:\n{str(e)}", parent=dialog)

        # Кнопки
        btn_frame = ctk.CTkFrame(form, fg_color="transparent")
        btn_frame.pack(fill="x", pady=20)

        ctk.CTkButton(
            btn_frame, text="Створити", command=do_create,
            font=ctk.CTkFont(size=13, weight="bold"),
            fg_color=_CLR_TEAL, hover_color=_CLR_TEAL_HOVER, height=40
        ).pack(side="left", padx=(0, 10))

        ctk.CTkButton(
            btn_frame, text="Скасувати", command=dialog.destroy,
            font=ctk.CTkFont(size=12),
            fg_color=_CLR_GRAY, hover_color=_CLR_GRAY_HOVER, height=40
        ).pack(side="right")

    # ==================== ЕКРАН РОЛЕЙ ====================

    def _show_roles_screen(self):
        self.current_screen = "roles"
        self._clear_screen()

        content = ctk.CTkFrame(self.root, fg_color="transparent")
        content.pack(fill="both", expand=True)

        self._make_header(content, "Ролі для БР")

        # Панель дій
        actions = ctk.CTkFrame(content, fg_color="transparent")
        actions.pack(fill="x", padx=15, pady=(0, 5))

        ctk.CTkLabel(actions, text="Місяць:", font=ctk.CTkFont(size=11)).pack(side="left", padx=(0, 5))

        self.roles_month_var = tk.StringVar()
        try:
            months = get_available_months(self.excel_file)
        except Exception:
            months = []
        months_combo = ctk.CTkComboBox(actions, variable=self.roles_month_var, values=months,
                                        width=220, state="readonly",
                                        command=lambda _: self._refresh_roles_treeview())
        months_combo.pack(side="left", padx=5)
        if months:
            months_combo.set(months[-1])

        ctk.CTkButton(
            actions, text="📥 Імпорт з табеля", command=self._import_from_tabel_action,
            font=ctk.CTkFont(size=11, weight="bold"),
            fg_color=_CLR_BLUE, hover_color=_CLR_BLUE_HOVER, height=34, width=160
        ).pack(side="left", padx=8)

        ctk.CTkButton(
            actions, text="⚙ Автопризначити", command=self._auto_assign_roles_action,
            font=ctk.CTkFont(size=11, weight="bold"),
            fg_color=_CLR_ORANGE, hover_color=_CLR_ORANGE_HOVER, height=34, width=160
        ).pack(side="left", padx=5)

        ctk.CTkButton(
            actions, text="← Назад", command=self._create_main_menu,
            font=ctk.CTkFont(size=11),
            fg_color=_CLR_GRAY, hover_color=_CLR_GRAY_HOVER, height=34, width=90
        ).pack(side="right", padx=5)

        # Treeview (ttk — CTk не має аналога)
        tree_frame = ctk.CTkFrame(content, fg_color="transparent")
        tree_frame.pack(fill="both", expand=True, padx=15, pady=5)

        columns = ("num", "pib", "rank", "position", "role")
        self.roles_tree = ttk.Treeview(
            tree_frame, columns=columns, show="headings", height=18,
            style="Dark.Treeview"
        )
        self.roles_tree.heading("num", text="#")
        self.roles_tree.heading("pib", text="ПІБ")
        self.roles_tree.heading("rank", text="Звання")
        self.roles_tree.heading("position", text="Посада")
        self.roles_tree.heading("role", text="Роль")

        self.roles_tree.column("num", width=40, stretch=False)
        self.roles_tree.column("pib", width=220)
        self.roles_tree.column("rank", width=120)
        self.roles_tree.column("position", width=180)
        self.roles_tree.column("role", width=200)

        scrollbar = ttk.Scrollbar(tree_frame, orient="vertical", command=self.roles_tree.yview)
        self.roles_tree.configure(yscrollcommand=scrollbar.set)

        self.roles_tree.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        self.roles_tree.bind("<<TreeviewSelect>>", self._on_role_row_select)

        self._roles_list = get_all_roles()

        # Панель редагування ролі
        edit_frame = ctk.CTkFrame(content)
        edit_frame.pack(fill="x", padx=15, pady=(5, 5))

        ctk.CTkLabel(edit_frame, text="Обраний:", font=ctk.CTkFont(size=11)).pack(side="left", padx=(10, 5))
        self._selected_pib_label = ctk.CTkLabel(
            edit_frame, text="— оберіть рядок —",
            font=ctk.CTkFont(size=11, weight="bold"), width=250, anchor="w"
        )
        self._selected_pib_label.pack(side="left", padx=(0, 15))

        ctk.CTkLabel(edit_frame, text="Роль:", font=ctk.CTkFont(size=11)).pack(side="left", padx=(0, 5))
        role_names = ["— не призначено —"] + [name for _, name in self._roles_list]
        self._role_combo = ctk.CTkComboBox(
            edit_frame, values=role_names, width=250, state="readonly",
            command=self._on_role_combo_changed
        )
        self._role_combo.pack(side="left", padx=(0, 10), pady=8)
        self._role_combo.set("— не призначено —")
        self._selected_role_pib = None  # ПІБ поточного вибраного рядка

        # Лог
        log_frame = ctk.CTkFrame(content, fg_color="transparent")
        log_frame.pack(fill="x", padx=15, pady=(0, 5))
        ctk.CTkLabel(log_frame, text="Статус", font=ctk.CTkFont(size=11, weight="bold")).pack(anchor="w")
        self.log_text = ctk.CTkTextbox(log_frame, height=80, font=ctk.CTkFont(family="Consolas", size=10))
        self.log_text.pack(fill="x")

        self._make_status_bar(content, "Оберіть рядок і змініть роль")

        self._refresh_roles_treeview()

    def _refresh_roles_treeview(self):
        if not hasattr(self, 'roles_tree'):
            return
        for item in self.roles_tree.get_children():
            self.roles_tree.delete(item)

        month = self.roles_month_var.get() if hasattr(self, 'roles_month_var') else ""
        if month:
            try:
                personnel = get_active_personnel_for_month(self.excel_file, month)
            except Exception:
                personnel = get_all_personnel()
        else:
            personnel = get_all_personnel()

        for i, p in enumerate(personnel, 1):
            role_display = p["role_name"] or "— не призначено —"
            self.roles_tree.insert("", "end", iid=p["pib"], values=(
                i, p["pib"], p["rank"], p["position"], role_display
            ))

    def _on_role_row_select(self, event):
        """При виборі рядка в Treeview — оновлює панель редагування ролі."""
        selection = self.roles_tree.selection()
        if not selection:
            return

        item_id = selection[0]
        values = self.roles_tree.item(item_id, "values")
        if not values or len(values) < 5:
            return

        pib = values[1]
        current_role = values[4]

        self._selected_role_pib = item_id  # iid = pib
        self._selected_pib_label.configure(text=pib)

        # Оновлюємо combobox без тригеру команди
        self._role_combo.set(current_role)

    def _on_role_combo_changed(self, selected_role: str):
        """При зміні ролі в combobox — зберігає в БД."""
        pib = self._selected_role_pib
        if not pib:
            return

        if selected_role == "— не призначено —":
            set_personnel_role(pib, None)
        else:
            role_id = None
            for rid, rname in self._roles_list:
                if rname == selected_role:
                    role_id = rid
                    break
            if role_id is not None:
                set_personnel_role(pib, role_id)

        self._refresh_roles_treeview()
        # Повертаємо виділення на той самий рядок
        if pib in [self.roles_tree.item(i, "values")[1] for i in self.roles_tree.get_children()]:
            self.roles_tree.selection_set(pib)

    def _import_from_tabel_action(self):
        month = self.roles_month_var.get()
        if not month:
            messagebox.showwarning("Увага", "Оберіть місяць!")
            return

        self._log(f"Імпорт з аркуша '{month}'...")
        self._update_status("Імпорт...")

        def do_import():
            try:
                count = import_personnel_from_tabel(self.excel_file, month)
                self.root.after(0, lambda: self._log(f"Імпортовано {count} записів."))
                self.root.after(0, self._refresh_roles_treeview)
                self.root.after(0, lambda: self._update_status(f"Імпорт завершено: {count} записів"))
            except Exception as e:
                self.root.after(0, lambda: self._log(f"ПОМИЛКА: {e}"))
                self.root.after(0, lambda: self._update_status("Помилка імпорту"))

        threading.Thread(target=do_import, daemon=True).start()

    def _auto_assign_roles_action(self):
        self._log("Автопризначення ролей...")
        self._update_status("Автопризначення...")

        def do_assign():
            try:
                stats = auto_assign_all_roles()
                total = sum(stats.values())
                self.root.after(0, lambda: self._log(f"Автопризначено {total} ролей:"))
                for role_name, count in stats.items():
                    self.root.after(0, lambda rn=role_name, c=count: self._log(f"  {rn}: {c}"))
                if not stats:
                    self.root.after(0, lambda: self._log("  Немає нових призначень."))
                self.root.after(0, self._refresh_roles_treeview)
                self.root.after(0, lambda: self._update_status(f"Автопризначено: {total}"))
            except Exception as e:
                self.root.after(0, lambda: self._log(f"ПОМИЛКА: {e}"))
                self.root.after(0, lambda: self._update_status("Помилка"))

        threading.Thread(target=do_assign, daemon=True).start()

    # ==================== ЗАГАЛЬНІ МЕТОДИ ====================

    def _get_busy_photo(self):
        if self._busy_photo:
            return self._busy_photo
        if not PIL_AVAILABLE:
            return None
        busy_path = os.path.join(get_base_path(), "alvares_busy.png")
        if not os.path.exists(busy_path):
            return None
        try:
            img = Image.open(busy_path).resize((24, 24), Image.Resampling.LANCZOS)
            self._busy_photo = ImageTk.PhotoImage(img)
            return self._busy_photo
        except Exception:
            return None

    def _make_status_bar(self, parent, text: str):
        status_frame = ctk.CTkFrame(parent, height=28, fg_color="transparent")
        status_frame.pack(side="bottom", fill="x", padx=10, pady=(0, 4))

        self._busy_icon_label = tk.Label(status_frame, bg=self.root.cget("bg"))
        self._busy_icon_label.pack(side="left", padx=(0, 4))
        self._busy_icon_label.pack_forget()

        self.status_label = ctk.CTkLabel(
            status_frame, text=text,
            font=ctk.CTkFont(size=10), text_color=_CLR_DIM,
            height=24, anchor="w"
        )
        self.status_label.pack(side="left", fill="x", expand=True)

    def _check_files(self):
        if not os.path.exists(self.excel_file):
            messagebox.showerror(
                "Помилка",
                f"Не знайдено файл: {self.excel_file}\n\n"
                "Переконайтеся, що файл знаходиться в поточній папці."
            )
        else:
            try:
                self.generator = ReportGenerator(self.excel_file)
            except Exception as e:
                messagebox.showerror("Помилка", f"Не вдалося завантажити файл:\n{str(e)}")

    def _log(self, message):
        if hasattr(self, 'log_text'):
            self.log_text.configure(state="normal")
            self.log_text.insert("end", message + "\n")
            self.log_text.see("end")
            self.log_text.configure(state="normal")
            self.root.update_idletasks()

    def _update_status(self, message):
        if hasattr(self, 'status_label'):
            self.status_label.configure(text=message)
            busy_keywords = ("Імпорт", "Автопризначен", "Генерація", "Заповнення", "Склад", "Створ")
            is_busy = any(message.startswith(kw) for kw in busy_keywords)
            if hasattr(self, '_busy_icon_label'):
                if is_busy:
                    photo = self._get_busy_photo()
                    if photo:
                        self._busy_icon_label.configure(image=photo)
                        self._busy_icon_label.image = photo
                        self._busy_icon_label.pack(side="left", padx=(0, 4))
                else:
                    self._busy_icon_label.pack_forget()
            self.root.update_idletasks()

    # ==================== ГЕНЕРАЦІЯ РАПОРТІВ ====================

    def _generate_reports(self):
        if not self.selected_month.get():
            messagebox.showwarning("Увага", "Оберіть місяць!")
            return

        month = self.selected_month.get()
        report_type = self.report_type_var.get()

        report_name = {
            "1": "ДГВ 100к", "2": "ДГВ 30к", "3": "Підтвердження 100к",
            "4": "Підтвердження 30к", "5": "ДГВ 0к", "6": "Всі типи рапортів"
        }.get(report_type, "Рапорт")

        if not messagebox.askyesno("Підтвердження", f"Створити {report_name} за {month}?"):
            return

        self.generate_btn.configure(state="disabled", text="⏳ Генерація...")
        self.log_text.configure(state="normal")
        self.log_text.delete("0.0", "end")

        thread = threading.Thread(target=self._do_generate_reports, args=(month, report_type))
        thread.daemon = True
        thread.start()

    def _do_generate_reports(self, month, report_type):
        try:
            self._update_status("Генерація рапортів...")
            self._log("=" * 60)
            self._log(f"Вітя Альварес розпочав генерацію даних за {month}...")
            self._log("=" * 60)

            self.generator.reader.load_workbook()
            soldiers = self.generator.reader.read_month_data(month)

            if not soldiers:
                self._log("❌ Не знайдено даних для цього місяця")
                self._update_status("Помилка: дані не знайдено")
                return

            self._log(f"✓ Знайдено та оброблено {len(soldiers)} військовослужбовців")

            month_display = month.replace("_", " ").lower()

            if report_type == "6":
                self._generate_all_reports(soldiers, month_display)
            else:
                self.generator._generate_report(month, report_type)
                self._log("✓ Рапорт успішно створено!")

            self._log("=" * 60)
            self._log("✓ Вітя Альварес роботу завершив — дані успішно створено!")
            self._log("=" * 60)
            self._update_status("Готово!")

            self.root.after(0, lambda: messagebox.showinfo(
                "Успіх", f"Рапорти успішно створено!\n\nПеревірте файли в поточній папці."
            ))

        except Exception as e:
            error_msg = f"Помилка: {str(e)}"
            self._log(f"❌ {error_msg}")
            self._update_status("Помилка!")
            self.root.after(0, lambda: messagebox.showerror("Помилка", error_msg))
        finally:
            self.root.after(0, lambda: self.generate_btn.configure(
                state="normal", text="📄 Створити рапорти"
            ))

    def _generate_all_reports(self, soldiers, month_display):
        self._log('"Працюю, як завжди швидко" © Вітя Альварес\n')

        reports = []

        soldiers_100 = self.generator.reader.get_soldiers_by_category(soldiers, "100", include_no_payment=False)
        if soldiers_100:
            filename = f"ДГВ_100к_{month_display}.xlsx"
            self.generator.excel_generator.create_dgv_report(soldiers_100, month_display, "100", filename)
            reports.append(filename)
            self._log(f"✓ Створено: {filename}")

        soldiers_30 = self.generator.reader.get_soldiers_by_category(soldiers, "30", include_no_payment=False)
        if soldiers_30:
            filename = f"ДГВ_30к_{month_display}.xlsx"
            self.generator.excel_generator.create_dgv_report(soldiers_30, month_display, "30", filename)
            reports.append(filename)
            self._log(f"✓ Створено: {filename}")

        soldiers_100_all = self.generator.reader.get_soldiers_by_category(soldiers, "100", include_no_payment=True)
        if soldiers_100_all:
            filename = f"Підтвердження_100к_{month_display}.docx"
            self.generator.word_generator.create_confirmation_report(soldiers_100_all, month_display, "100", filename)
            reports.append(filename)
            self._log(f"✓ Створено: {filename}")

        soldiers_30_all = self.generator.reader.get_soldiers_by_category(soldiers, "30", include_no_payment=True)
        if soldiers_30_all:
            filename = f"Підтвердження_30к_{month_display}.docx"
            self.generator.word_generator.create_confirmation_report(soldiers_30_all, month_display, "30", filename)
            reports.append(filename)
            self._log(f"✓ Створено: {filename}")

        soldiers_0 = self.generator.reader.get_soldiers_by_category(soldiers, "0", include_no_payment=False)
        if soldiers_0:
            filename = f"ДГВ_0к_{month_display}.xlsx"
            self.generator.excel_generator.create_dgv_report(soldiers_0, month_display, "0", filename)
            reports.append(filename)
            self._log(f"✓ Створено: {filename}")

        self._log(f"\n✓ Всього створено файлів: {len(reports)}")

    # ==================== ЗАПОВНЕННЯ ТАБЕЛЮ ====================

    def _fill_tabel(self):
        selected = self.tabel_month_var.get()
        if not selected:
            messagebox.showwarning("Увага", "Оберіть місяць!")
            return

        if not messagebox.askyesno("Підтвердження", f"Заповнити табель за {selected}?"):
            return

        self.fill_tabel_btn.configure(state="disabled", text="⏳ Заповнення...")
        self.log_text.configure(state="normal")
        self.log_text.delete("0.0", "end")

        thread = threading.Thread(target=self._do_fill_tabel, args=(selected,))
        thread.daemon = True
        thread.start()

    def _do_fill_tabel(self, selected_month):
        try:
            self._update_status("Заповнення табелю...")
            self._log("=" * 60)
            self._log(f"Заповнення табелю: {selected_month}")
            self._log("=" * 60)

            import sys
            from io import StringIO
            old_stdout = sys.stdout
            sys.stdout = StringIO()

            try:
                if selected_month == "-- Всі місяці --":
                    fill_tabel_months(self.excel_file)
                else:
                    parsed = parse_month_sheet_name(selected_month)
                    if not parsed:
                        raise ValueError(f"Не вдалося розпарсити назву: {selected_month}")
                    year, month_num = parsed
                    source_file = get_source_filename(selected_month)
                    source_path = os.path.join(get_app_dir(), source_file)
                    fill_single_month(selected_month, source_path, year, month_num, self.excel_file)

                output = sys.stdout.getvalue()
                self._log(output)
            finally:
                sys.stdout = old_stdout

            self._log("=" * 60)
            self._log("Заповнення завершено!")
            self._update_status("Готово!")

            self.root.after(0, lambda: messagebox.showinfo(
                "Успіх", "Табель успішно заповнено!"
            ))

        except Exception as e:
            error_msg = f"Помилка: {str(e)}"
            self._log(f"{error_msg}")
            self._update_status("Помилка!")
            self.root.after(0, lambda: messagebox.showerror("Помилка", error_msg))
        finally:
            self.root.after(0, lambda: self.fill_tabel_btn.configure(
                state="normal", text="📊 Заповнити табель"
            ))

    # ==================== WORD БР ====================

    def _preview_composition(self):
        date_str = self.start_date_var.get().strip()
        if not date_str:
            messagebox.showwarning("Увага", "Введіть дату БР!")
            return

        try:
            br_date = datetime.strptime(date_str, "%d.%m.%Y")
        except ValueError:
            messagebox.showerror("Помилка", "Некоректний формат дати (ДД.ММ.РРРР)")
            return

        self.log_text.configure(state="normal")
        self.log_text.delete("0.0", "end")
        self._log(f"Формування складу на дату БР: {date_str}")
        from datetime import timedelta as _td
        tabel_date = br_date + _td(days=1)
        self._log(f"Дата табеля (БР+1): {tabel_date.strftime('%d.%m.%Y')}")
        self._log("")

        def do_preview():
            try:
                composition = build_composition_for_date(self.excel_file, br_date)
                total = 0
                for role_name, members in composition.items():
                    count = len(members)
                    total += count
                    self.root.after(0, lambda rn=role_name, c=count: self._log(f"--- {rn} ({c}) ---"))
                    if members:
                        for m in members:
                            self.root.after(0, lambda mm=m: self._log(f"  {mm['rank']} {mm['pib']}"))
                    else:
                        self.root.after(0, lambda: self._log("  (порожньо)"))
                    self.root.after(0, lambda: self._log(""))
                self.root.after(0, lambda: self._log(f"Всього: {total} осіб з відміткою 100"))
                self.root.after(0, lambda: self._update_status(f"Склад сформовано: {total} осіб"))
            except Exception as e:
                self.root.after(0, lambda: self._log(f"ПОМИЛКА: {e}"))
                self.root.after(0, lambda: self._update_status("Помилка"))

        threading.Thread(target=do_preview, daemon=True).start()

    def _generate_word_br(self):
        start_str = self.start_date_var.get().strip()
        end_str = self.end_date_var.get().strip()
        if not start_str or not end_str:
            messagebox.showwarning("Увага", "Введіть початкову та кінцеву дату!")
            return

        try:
            start_date = datetime.strptime(start_str, "%d.%m.%Y")
            end_date = datetime.strptime(end_str, "%d.%m.%Y")
        except ValueError:
            messagebox.showerror("Помилка", "Некоректний формат дати (ДД.ММ.РРРР)")
            return

        if start_date > end_date:
            messagebox.showerror("Помилка", "Початкова дата не може бути пізнішою за кінцеву!")
            return

        if not os.path.exists(self.template_path):
            messagebox.showerror("Помилка", f"Шаблон не знайдено: {self.template_path}")
            return

        self.log_text.configure(state="normal")
        self.log_text.delete("0.0", "end")
        self._log(f"Генерація Word БР: {start_str} — {end_str}")

        def do_generate():
            from datetime import timedelta
            try:
                created = 0
                current = start_date
                while current <= end_date:
                    ds = current.strftime("%d.%m.%Y")
                    self.root.after(0, lambda d=ds: self._log(f"\n--- БР на {d} ---"))
                    composition = build_composition_for_date(self.excel_file, current)
                    total = sum(len(m) for m in composition.values())
                    self.root.after(0, lambda t=total: self._log(f"  Осіб з роллю: {t}"))

                    result_path = generate_br_word(
                        current, composition, self.template_path, self.output_dir,
                        br_4shb_file=self.br_4shb_file,
                        tabel_file=self.excel_file
                    )
                    self.root.after(0, lambda p=result_path: self._log(f"  Створено: {p}"))
                    created += 1

                    # Генерація БР для бійців з першим днем "роп"
                    rop_path = generate_rop_word(
                        current, self.excel_file, self.rop_template_path,
                        self.output_dir, br_4shb_file=self.br_4shb_file
                    )
                    if rop_path:
                        self.root.after(0, lambda p=rop_path: self._log(f"  РОП: {p}"))
                        created += 1

                    current += timedelta(days=1)

                self.root.after(0, lambda: self._log(f"\nВсього створено {created} файлів БР"))
                self.root.after(0, lambda: self._update_status(f"Створено {created} БР"))
                self.root.after(0, lambda: messagebox.showinfo(
                    "Готово", f"Створено {created} файлів БР"
                ))
            except Exception as e:
                self.root.after(0, lambda: self._log(f"ПОМИЛКА: {e}"))
                self.root.after(0, lambda: self._update_status("Помилка генерації"))

        threading.Thread(target=do_generate, daemon=True).start()


def main():
    """Запускає графічний інтерфейс"""
    root = ctk.CTk()
    app = ReportGUI(root)
    root.mainloop()


if __name__ == "__main__":
    main()
