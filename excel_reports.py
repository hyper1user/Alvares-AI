"""
Модуль генерації Excel-рапортів (ДГВ 100к/30к/0к)
"""
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from datetime import datetime
from typing import List
from excel_processor import SoldierData

class ExcelReportGenerator:
    """Клас для генерації Excel-рапортів"""
    
    def __init__(self):
        self.header_font = Font(bold=True, size=12)
        self.cell_font = Font(size=10)
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.center_alignment = Alignment(horizontal='center', vertical='center')
        self.left_alignment = Alignment(horizontal='left', vertical='center')
    
    def create_dgv_report(self, soldiers: List[SoldierData], month_name: str, 
                        category: str, output_file: str) -> str:
        """
        Створює рапорт на ДГВ
        
        Args:
            soldiers: Список бійців
            month_name: Назва місяця
            category: "100", "30", або "0"
            output_file: Шлях до файлу
            
        Returns:
            str: Шлях до створеного файлу
        """
        
        wb = Workbook()
        ws = wb.active
        ws.title = f"ДГВ_{category}к"
        
        # Заголовок документа
        self._add_header(ws, month_name, category)
        
        # Таблиця з даними
        self._add_soldiers_table(ws, soldiers, category)
        
        # Налаштування стовпців
        self._adjust_columns(ws)
        
        # Зберігаємо файл
        wb.save(output_file)
        print(f"Створено ДГВ: {output_file}")
        return output_file
    
    def _add_header(self, ws, month_name: str, category: str):
        """Додає заголовок документа"""
        
        # Основний заголовок
        ws['C1'] = "ВІДОМІСТЬ"
        ws['C1'].font = Font(bold=True, size=14)
        ws['C1'].alignment = self.center_alignment
        
        ws['C2'] = "про участь військовослужбовців 12ШР 4ШБ у бойових діях"
        ws['C2'].font = self.header_font
        ws['C2'].alignment = self.center_alignment
        
        # Назва підрозділу та період
        period_text = f"за {month_name} місяць"
        ws['C3'] = period_text
        ws['C3'].font = self.header_font
        ws['C3'].alignment = self.center_alignment
        
        # Пояснення категорії
        ws['C5'] = self._get_category_explanation(category)
        ws['C5'].font = self.cell_font
        
        # Порожній рядок
        ws['C6'] = ""
    
    def _get_category_explanation(self, category: str) -> str:
        """Повертає пояснення категорії"""
        explanations = {
            "100": "1. Військовослужбовці, які безпосередньо брали участь у бойових діях",
            "30": "2. Військовослужбовці (забезпечуючі) військові частини в районі проведення бойових дій",
            "0": "3. Військовослужбовці, які не брали безпосередню участь у бойових діях"
        }
        return explanations.get(category, "")
    
    def _add_soldiers_table(self, ws, soldiers: List[SoldierData], category: str):
        """Додає таблицю з даними бійців"""
        
        # Заголовки таблиці
        headers = [
            "№ п/п",
            "Військове звання",
            "Прізвище ім'я по батькові",
            "Період участі",
            "Кількість днів",
            "Категорія нарахувань",
            "Примітка"
        ]
        
        # Записуємо заголовки
        for col, header in enumerate(headers, 1):
            cell = ws.cell(7, col, header)
            cell.font = self.header_font
            cell.alignment = self.center_alignment
            cell.border = self.thin_border
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Додаємо дані бійців
        for i, soldier in enumerate(soldiers, 1):
            row = 7 + i
            
            # № п/п
            ws.cell(row, 1, i).border = self.thin_border
            
            # Звання
            ws.cell(row, 2, soldier.rank).border = self.thin_border
            ws.cell(row, 2).alignment = self.left_alignment
            
            # ПІБ
            ws.cell(row, 3, soldier.pib).border = self.thin_border
            ws.cell(row, 3).alignment = self.left_alignment
            
            # Період
            period = self._get_period_for_category(soldier, category)
            ws.cell(row, 4, period).border = self.thin_border
            ws.cell(row, 4).alignment = self.left_alignment
            
            # Кількість днів
            days_count = self._get_days_count_for_category(soldier, category)
            ws.cell(row, 5, days_count).border = self.thin_border
            ws.cell(row, 5).alignment = self.center_alignment
            
            # Категорія
            amount = self._get_amount_for_category(category)
            ws.cell(row, 6, amount).border = self.thin_border
            ws.cell(row, 6).alignment = self.center_alignment
            
            # Примітка
            note = ""
            if soldier.has_no_payment_note():
                note = "не виплачувати"
            ws.cell(row, 7, note).border = self.thin_border
            ws.cell(row, 7).alignment = self.left_alignment
    
    def _get_period_for_category(self, soldier: SoldierData, category: str) -> str:
        """Повертає період для категорії"""
        if category == "100" and soldier.days_100:
            return self._format_period(soldier.days_100)
        elif category == "30" and soldier.days_30:
            return self._format_period(soldier.days_30)
        elif category == "0" and soldier.days_0:
            return self._format_period(soldier.days_0)
        return ""
    
    def _get_days_count_for_category(self, soldier: SoldierData, category: str) -> int:
        """Повертає кількість днів для категорії"""
        if category == "100":
            return len(soldier.days_100)
        elif category == "30":
            return len(soldier.days_30)
        elif category == "0":
            return len(soldier.days_0)
        return 0
    
    def _get_amount_for_category(self, category: str) -> str:
        """Повертає суму для категорії"""
        amounts = {
            "100": "100 000",
            "30": "30 000",
            "0": "0"
        }
        return amounts.get(category, "0")
    
    def _format_period(self, dates: List[datetime]) -> str:
        """Форматує період"""
        if not dates:
            return ""
        
        dates.sort()
        start_date = dates[0]
        end_date = dates[-1]
        
        if start_date == end_date:
            return start_date.strftime("%d.%m.%Y")
        else:
            return f"з {start_date.strftime('%d.%m.%Y')} по {end_date.strftime('%d.%m.%Y')}"
    
    def _adjust_columns(self, ws):
        """Налаштовує ширину стовпців"""
        column_widths = {
            'A': 8,   # № п/п
            'B': 20,  # Звання
            'C': 30,  # ПІБ
            'D': 25,  # Період
            'E': 12,  # Кількість днів
            'F': 20,  # Категорія
            'G': 15   # Примітка
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

def create_sample_excel_reports():
    """Створює зразки Excel-рапортів для тестування"""
    
    from excel_processor import TabelReader
    
    # Читаємо дані
    reader = TabelReader("Табель_Багатомісячний.xlsx")
    soldiers = reader.read_month_data("Листопад_2025")
    
    # Фільтруємо бійців
    soldiers_100 = reader.get_soldiers_by_category(soldiers, "100", include_no_payment=False)
    soldiers_30 = reader.get_soldiers_by_category(soldiers, "30", include_no_payment=False)
    soldiers_0 = reader.get_soldiers_by_category(soldiers, "0", include_no_payment=False)
    
    # Генеруємо рапорти
    generator = ExcelReportGenerator()
    
    # ДГВ 100к
    if soldiers_100:
        generator.create_dgv_report(
            soldiers_100, 
            "листопад 2025", 
            "100", 
            "ДГВ_100к_листопад_2025.xlsx"
        )
    
    # ДГВ 30к
    if soldiers_30:
        generator.create_dgv_report(
            soldiers_30, 
            "листопад 2025", 
            "30", 
            "ДГВ_30к_листопад_2025.xlsx"
        )
    
    # ДГВ 0к
    if soldiers_0:
        generator.create_dgv_report(
            soldiers_0, 
            "листопад 2025", 
            "0", 
            "0к_листопад_2025.xlsx"
        )

if __name__ == "__main__":
    create_sample_excel_reports()

