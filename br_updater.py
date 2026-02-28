"""
Утиліти для роботи з БР документами: парсинг дат, форматування ПІБ, читання табеля.
"""
import re
import os
from datetime import datetime, timedelta
from month_utils import parse_month_sheet_name
import openpyxl
from typing import List, Tuple

# Кеш workbook: {шлях_файлу: workbook} — щоб не відкривати файл по 5 разів на кожен день
_wb_cache = {}


def _get_workbook(tabel_file: str):
    """Повертає кешований workbook або завантажує новий."""
    real_path = os.path.realpath(tabel_file)
    if real_path not in _wb_cache:
        _wb_cache[real_path] = openpyxl.load_workbook(tabel_file, data_only=True)
    return _wb_cache[real_path]


def clear_wb_cache():
    """Очищає кеш workbook (викликати після завершення генерації)."""
    _wb_cache.clear()


def parse_filename_date(filename: str) -> datetime:
    """
    Парсить дату з назви БР файлу

    Формат: БР_№_121_30.04.2025.docx -> datetime(2025, 4, 30)

    Args:
        filename: Назва файлу

    Returns:
        datetime: Дата БР
    """
    # Шукаємо паттерн дати dd.mm.yyyy
    match = re.search(r'(\d{2})\.(\d{2})\.(\d{4})', filename)
    if not match:
        raise ValueError(f"Не вдалося знайти дату в назві файлу: {filename}")

    day, month, year = map(int, match.groups())
    return datetime(year, month, day)


def get_tabel_date(br_date: datetime) -> datetime:
    """
    Повертає дату для табелю (+1 день від дати БР)

    Args:
        br_date: Дата БР

    Returns:
        datetime: Дата для табелю
    """
    return br_date + timedelta(days=1)


def pib_to_document_format(pib: str, rank: str = "") -> str:
    """
    Конвертує ПІБ у формат для документа БР з додаванням звання
    Формат: "Прізвище Ім'я По-батькові" -> "ПРІЗВИЩЕ Ім'я По-батькові"
    З додаванням звання: "звання ПРІЗВИЩЕ Ім'я По-батькові"

    Args:
        pib: ПІБ у форматі з табелю
        rank: Військове звання (необов'язково)

    Returns:
        str: ПІБ у форматі для БР документа з званням
    """
    parts = pib.strip().split()
    if len(parts) < 2:
        pib_formatted = pib.upper()
    else:
        # Прізвище великими, решта як є
        surname = parts[0].upper()
        rest = ' '.join(parts[1:])
        pib_formatted = f"{surname} {rest}"

    # Додаємо звання на початок, якщо воно вказане
    if rank and rank.strip():
        return f"{rank.strip()} {pib_formatted}"
    return pib_formatted


def pib_to_table_format(pib: str, rank: str = "") -> str:
    """
    Конвертує ПІБ у формат для таблиці вкінці БР з додаванням звання
    Формат: "Прізвище Ім'я По-батькові" -> "Ім'я ПРІЗВИЩЕ"
    З додаванням звання: "звання Ім'я ПРІЗВИЩЕ"

    Args:
        pib: ПІБ у форматі з табелю
        rank: Військове звання (необов'язково)

    Returns:
        str: ПІБ у форматі для таблиці з званням
    """
    parts = pib.strip().split()
    if len(parts) < 2:
        pib_formatted = pib.upper()
    else:
        surname = parts[0].upper()
        name = parts[1]
        pib_formatted = f"{name} {surname}"

    # Додаємо звання на початок, якщо воно вказане
    if rank and rank.strip():
        return f"{rank.strip()} {pib_formatted}"
    return pib_formatted


def normalize_pib(pib: str) -> str:
    """
    Нормалізує ПІБ для порівняння (видаляє зайві пробіли)

    Args:
        pib: ПІБ

    Returns:
        str: Нормалізований ПІБ
    """
    return ' '.join(pib.strip().split())


def get_soldiers_from_tabel(tabel_file: str, date: datetime) -> Tuple[List[Tuple[str, str]], List[Tuple[str, str]]]:
    """
    Отримує списки ПІБ та звань військовослужбовців з позначками 100 та 30 для заданої дати.
    Бійці з позначкою "роп" включаються до списку 100 (для табелю/виплат).

    Args:
        tabel_file: Шлях до файлу табелю
        date: Дата

    Returns:
        Tuple[List[Tuple[str, str]], List[Tuple[str, str]]]:
            (список (ПІБ, звання) з позначкою 100 або роп, список (ПІБ, звання) з позначкою 30)
    """
    soldiers_100, soldiers_rop, soldiers_30 = _get_soldiers_from_tabel_detailed(tabel_file, date)
    # Для зворотної сумісності: 100 + роп разом
    return soldiers_100 + soldiers_rop, soldiers_30


def _get_soldiers_from_tabel_detailed(
    tabel_file: str, date: datetime
) -> Tuple[List[Tuple[str, str]], List[Tuple[str, str]], List[Tuple[str, str]]]:
    """
    Отримує списки ПІБ та звань з розділенням 100, роп та 30.

    Returns:
        (soldiers_100, soldiers_rop, soldiers_30)
    """
    wb = _get_workbook(tabel_file)

    sheet_name = None
    for name in wb.sheetnames:
        parsed = parse_month_sheet_name(name)
        if parsed and parsed[0] == date.year and parsed[1] == date.month:
            sheet_name = name
            break

    if not sheet_name:
        raise ValueError(f"Аркуш для {date.strftime('%m.%Y')} не знайдено")

    ws = wb[sheet_name]

    header_row = None
    for row in range(1, 20):
        if ws.cell(row, 6).value and "ПІБ" in str(ws.cell(row, 6).value):
            header_row = row
            break

    if not header_row:
        raise ValueError("Не знайдено рядок з заголовками")

    col_num = 6 + date.day

    soldiers_100 = []
    soldiers_rop = []
    soldiers_30 = []

    for row in range(header_row + 1, ws.max_row + 1):
        pib = ws.cell(row, 6).value
        rank = ws.cell(row, 5).value or ""

        if not pib or not str(pib).strip():
            continue

        pib_str = str(pib).strip()
        rank_str = str(rank).strip() if rank else ""
        mark = ws.cell(row, col_num).value

        if mark:
            mark_str = str(mark).strip()
            # openpyxl може повертати float (100.0 замість 100)
            try:
                num_val = float(mark_str)
                if num_val == int(num_val):
                    mark_str = str(int(num_val))
            except (ValueError, TypeError):
                pass
            mark_str = mark_str.lower()
            if mark_str == "100":
                soldiers_100.append((pib_str, rank_str))
            elif mark_str == "роп":
                soldiers_rop.append((pib_str, rank_str))
            elif mark_str == "30":
                soldiers_30.append((pib_str, rank_str))

    return soldiers_100, soldiers_rop, soldiers_30


def get_first_rop_entries(
    tabel_file: str, br_date: datetime
) -> List[Tuple[str, str, str]]:
    """
    Знаходить бійців, у яких на дату табеля (br_date+1) починається серія 'роп'.
    Тобто: tabel_day == 'роп' AND попередній день != 'роп'.

    Returns:
        [(pib, rank, position)] — позиція з маленької літери
    """
    tabel_date = get_tabel_date(br_date)

    wb = _get_workbook(tabel_file)

    sheet_name = None
    for name in wb.sheetnames:
        parsed = parse_month_sheet_name(name)
        if parsed and parsed[0] == tabel_date.year and parsed[1] == tabel_date.month:
            sheet_name = name
            break

    if not sheet_name:
        return []

    ws = wb[sheet_name]

    header_row = None
    for row in range(1, 20):
        if ws.cell(row, 6).value and "ПІБ" in str(ws.cell(row, 6).value):
            header_row = row
            break

    if not header_row:
        return []

    tabel_col = 6 + tabel_date.day
    prev_col = 6 + br_date.day if br_date.month == tabel_date.month else None

    result = []
    for row in range(header_row + 1, ws.max_row + 1):
        pib = ws.cell(row, 6).value
        if not pib or not str(pib).strip():
            continue

        mark = ws.cell(row, tabel_col).value
        if not mark or str(mark).strip().lower() != "роп":
            continue

        # Перевіряємо попередній день
        if prev_col:
            prev_mark = ws.cell(row, prev_col).value
            prev_str = str(prev_mark).strip().lower() if prev_mark else ""
            if prev_str == "роп":
                continue  # Не перший день серії
        # Якщо prev_col is None (перший день місяця) — вважаємо першим роп

        pib_str = str(pib).strip()
        rank_str = str(ws.cell(row, 5).value or "").strip()
        position = ws.cell(row, 4).value
        position_str = str(position).strip() if position else ""
        if position_str:
            position_str = position_str[0].lower() + position_str[1:]

        result.append((pib_str, rank_str, position_str))

    return result


def get_soldiers_returning_from_rop(
    tabel_file: str, br_date: datetime
) -> List[Tuple[str, str]]:
    """
    Знаходить бійців, у яких серія 'роп' щойно закінчилась:
    - на дату БР (br_date) позначка == 'роп'
    - на дату табеля (br_date+1) позначка != 'роп' і != '100' (порожньо)
    Без них ці бійці взагалі не потрапляють до складу БР.

    Returns:
        [(pib, rank)]
    """
    tabel_date = get_tabel_date(br_date)

    try:
        _, rop_on_br_date, _ = _get_soldiers_from_tabel_detailed(tabel_file, br_date)
    except ValueError:
        return []

    if not rop_on_br_date:
        return []

    try:
        soldiers_100_tabel, rop_on_tabel_date, _ = _get_soldiers_from_tabel_detailed(tabel_file, tabel_date)
    except ValueError:
        return []

    # ПІБ тих, хто вже є в складі через "100" або ще "роп" на дату табеля
    already_in_br = {normalize_pib(p) for p, _ in soldiers_100_tabel + rop_on_tabel_date}

    return [
        (pib, rank)
        for pib, rank in rop_on_br_date
        if normalize_pib(pib) not in already_in_br
    ]


def get_continuing_rop_entries(
    tabel_file: str, br_date: datetime
) -> List[Tuple[str, str, str]]:
    """
    Знаходить бійців, у яких на дату табеля (br_date+1) продовжується серія 'роп'.
    Тобто: tabel_day == 'роп' AND попередній день == 'роп' (не перший день).

    Returns:
        [(pib, rank, position)] — позиція з маленької літери
    """
    tabel_date = get_tabel_date(br_date)

    wb = _get_workbook(tabel_file)

    sheet_name = None
    for name in wb.sheetnames:
        parsed = parse_month_sheet_name(name)
        if parsed and parsed[0] == tabel_date.year and parsed[1] == tabel_date.month:
            sheet_name = name
            break

    if not sheet_name:
        return []

    ws = wb[sheet_name]

    header_row = None
    for row in range(1, 20):
        if ws.cell(row, 6).value and "ПІБ" in str(ws.cell(row, 6).value):
            header_row = row
            break

    if not header_row:
        return []

    tabel_col = 6 + tabel_date.day
    prev_col = 6 + br_date.day if br_date.month == tabel_date.month else None

    result = []
    for row in range(header_row + 1, ws.max_row + 1):
        pib = ws.cell(row, 6).value
        if not pib or not str(pib).strip():
            continue

        mark = ws.cell(row, tabel_col).value
        if not mark or str(mark).strip().lower() != "роп":
            continue

        # Тільки ті, у кого попередній день теж "роп" (продовження серії)
        if not prev_col:
            continue  # Перший день місяця — вважаємо першим роп, пропускаємо
        prev_mark = ws.cell(row, prev_col).value
        prev_str = str(prev_mark).strip().lower() if prev_mark else ""
        if prev_str != "роп":
            continue  # Перший день серії — пропускаємо

        pib_str = str(pib).strip()
        rank_str = str(ws.cell(row, 5).value or "").strip()
        position = ws.cell(row, 4).value
        position_str = str(position).strip() if position else ""
        if position_str:
            position_str = position_str[0].lower() + position_str[1:]

        result.append((pib_str, rank_str, position_str))

    return result
