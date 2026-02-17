"""
Централізований модуль для роботи з місяцями.
Автоматично визначає доступні місяці з аркушів Табель_Багатомісячний.xlsx.
"""
import re
import openpyxl
from datetime import datetime
from typing import List, Tuple, Optional, Dict

# Патч для обходу помилки openpyxl з фільтрами таблиць
# (ValueError: Value must be either numerical or a string containing a wildcard)
try:
    from openpyxl.worksheet.filters import CustomFilterValueDescriptor
    _original_set = CustomFilterValueDescriptor.__set__
    def _patched_set(self, instance, value):
        try:
            _original_set(self, instance, value)
        except ValueError:
            instance.__dict__[self.name] = value
    CustomFilterValueDescriptor.__set__ = _patched_set
except Exception:
    pass

# Українські назви місяців -> номер
MONTH_NAMES_UK: Dict[str, int] = {
    "січень": 1, "лютий": 2, "березень": 3, "квітень": 4,
    "травень": 5, "червень": 6, "липень": 7,
    "серпень": 8, "вересень": 9, "жовтень": 10,
    "листопад": 11, "грудень": 12
}

# Номер -> українська назва (з великої літери)
MONTH_NAMES_UK_REVERSE: Dict[int, str] = {v: k.capitalize() for k, v in MONTH_NAMES_UK.items()}

# Номер -> українська назва (з маленької літери)
MONTH_NAMES_UK_LOWER: Dict[int, str] = {v: k for k, v in MONTH_NAMES_UK.items()}


def parse_month_sheet_name(sheet_name: str) -> Optional[Tuple[int, int]]:
    """
    Парсить назву аркуша у (рік, номер_місяця).

    "Січень_2026" -> (2026, 1)
    Повертає None якщо назва не відповідає патерну.
    """
    match = re.match(r'^([А-ЯІЄЇҐа-яієїґ]+)_(\d{4})$', sheet_name)
    if not match:
        return None
    month_name = match.group(1).lower()
    year = int(match.group(2))
    month_num = MONTH_NAMES_UK.get(month_name)
    if month_num is None:
        return None
    return (year, month_num)


def get_available_months(excel_file: str) -> List[str]:
    """
    Читає назви аркушів з Excel файлу та повертає ті,
    що відповідають патерну Місяць_Рік, відсортовані хронологічно.
    """
    wb = openpyxl.load_workbook(excel_file, read_only=True)
    months = []
    for name in wb.sheetnames:
        parsed = parse_month_sheet_name(name)
        if parsed:
            year, month_num = parsed
            months.append((year, month_num, name))
    wb.close()
    months.sort(key=lambda x: (x[0], x[1]))
    return [name for _, _, name in months]


def get_sheet_name_for_date(date: datetime, sheetnames: list) -> str:
    """
    Знаходить назву аркуша для заданої дати зі списку назв аркушів.

    Args:
        date: Дата для пошуку
        sheetnames: Список назв аркушів (wb.sheetnames)
    """
    for name in sheetnames:
        parsed = parse_month_sheet_name(name)
        if parsed and parsed[0] == date.year and parsed[1] == date.month:
            return name
    raise ValueError(f"Аркуш для {date.strftime('%m.%Y')} не знайдено")


def get_source_filename(sheet_name: str) -> str:
    """
    Повертає назву файлу-джерела для місяця.

    "Січень_2026" -> "Січень_2026.xlsx"
    """
    return f"{sheet_name}.xlsx"


def build_month_sheet_name(year: int, month: int) -> str:
    """
    Будує назву аркуша з року та номера місяця.

    (2026, 2) -> "Лютий_2026"
    """
    return f"{MONTH_NAMES_UK_REVERSE[month]}_{year}"
